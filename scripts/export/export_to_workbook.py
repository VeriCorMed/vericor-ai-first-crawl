# export_to_workbook.py
"""
Builds an Excel workbook (audit_export.xlsx) with three sheets:
- Products  (from pages_clean/products/*.md)
- Pages     (from pages_clean/pages/*.md)
- Posts     (from pages_clean/posts/*.md)

It reads YAML front-matter (if present) and the Markdown body.
If a file has no front-matter, metadata columns are left blank and
title is inferred from the first H1 in the body.

Run:
    python export_to_workbook.py
Optional:
    python export_to_workbook.py --out audit_export_YYYYMMDD.xlsx
"""

from __future__ import annotations
import argparse
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

# ---- dependencies: openpyxl + python-frontmatter ----
try:
    import frontmatter
except Exception as e:
    raise SystemExit(
        "Missing dependency 'python-frontmatter'. Install with:\n"
        "    pip install python-frontmatter\n"
        f"Details: {e}"
    )

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except Exception as e:
    raise SystemExit(
        "Missing dependency 'openpyxl'. Install with:\n"
        "    pip install openpyxl\n"
        f"Details: {e}"
    )

# ---------- paths ----------
ROOT = Path(__file__).resolve().parent
CLEAN_ROOT = ROOT / "pages_clean"
DIR_PRODUCTS = CLEAN_ROOT / "products"
DIR_PAGES = CLEAN_ROOT / "pages"
DIR_POSTS = CLEAN_ROOT / "posts"

# if your live site base is useful for building URLs when missing:
BASE_URL = "https://www.vericormed.com"

# ---------- small helpers ----------
def read_md(path: Path) -> Tuple[Dict[str, Any], str]:
    """Return (metadata, body) from a Markdown file with optional YAML front-matter."""
    post = frontmatter.load(path, encoding="utf-8")
    meta = dict(post.metadata or {})
    body = post.content or ""
    return meta, body

def infer_title_from_body(body: str) -> str:
    """Use the first ATX H1 (# Title) as title, else first non-empty line."""
    for line in body.splitlines():
        if line.strip().startswith("# "):
            return line.strip()[2:].strip()
    for line in body.splitlines():
        if line.strip():
            return line.strip()
    return ""

_md_link_re = re.compile(r"\[([^\]]+)\]\([^)]+\)")
_md_img_re = re.compile(r"!\[[^\]]*\]\([^)]+\)")

def markdown_to_plain_text(s: str) -> str:
    """Very light markdown cleanup for word count / excerpt."""
    # remove images entirely
    s = _md_img_re.sub(" ", s)
    # replace links [text](url) -> text
    s = _md_link_re.sub(r"\1", s)
    # remove code fences
    s = re.sub(r"```.*?```", " ", s, flags=re.DOTALL)
    # remove leftover markdown tokens
    s = re.sub(r"[#*_>`~|]", " ", s)
    # collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s

def word_count(text: str) -> int:
    text = text.strip()
    if not text:
        return 0
    return len(re.findall(r"\w+", text))

def first_n(text: str, n: int = 200) -> str:
    return text[:n] + ("…" if len(text) > n else "")

def ensure_sheet_headers(ws: Worksheet, headers: List[str]) -> None:
    ws.append(headers)
    # bold header and set some reasonable widths
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = min(max(12, len(header) + 2), 50)

# ---------- collectors ----------
def collect_products() -> Tuple[List[str], List[List[Any]]]:
    """
    Returns (headers, rows).
    Columns chosen to be stable even if some front-matter is missing.
    """
    headers = [
        "id", "sku", "title", "slug", "type", "status",
        "price", "regular_price", "sale_price",
        "stock_status", "stock_quantity",
        "weight", "dim_length", "dim_width", "dim_height",
        "shipping_class_name", "shipping_class_slug",
        "categories", "tags",
        "product_url",
        "images_count",
        "word_count",
        "excerpt",
        "file_name",
    ]
    rows: List[List[Any]] = []

    if not DIR_PRODUCTS.exists():
        return headers, rows

    files = sorted(DIR_PRODUCTS.glob("*.md"))
    for fp in files:
        meta, body = read_md(fp)

        # images may be a list of {src, alt, name, position}
        images = meta.get("images") or []
        if isinstance(images, list):
            img_count = len(images)
        else:
            img_count = 0

        # shipping_class may be {} or {'name':..,'slug':..}
        sc = meta.get("shipping_class") or {}
        sc_name = sc.get("name", "") if isinstance(sc, dict) else ""
        sc_slug = sc.get("slug", "") if isinstance(sc, dict) else ""

        # categories / tags may be lists (of strings)
        cats = meta.get("categories") or []
        if isinstance(cats, list):
            cats_str = ", ".join(str(c).strip() for c in cats)
        else:
            cats_str = str(cats)

        tags = meta.get("tags") or []
        if isinstance(tags, list):
            tags_str = ", ".join(str(t).strip() for t in tags)
        else:
            tags_str = str(tags)

        # dimensions
        dims = meta.get("dimensions") or {}
        d_len = (dims.get("length") if isinstance(dims, dict) else "") or ""
        d_wid = (dims.get("width") if isinstance(dims, dict) else "") or ""
        d_hgt = (dims.get("height") if isinstance(dims, dict) else "") or ""

        # basic text & excerpt
        plain = markdown_to_plain_text(body)
        wc = word_count(plain)
        excerpt = first_n(plain, 220)

        rows.append([
            meta.get("id", ""),
            meta.get("sku", ""),
            meta.get("title", "") or infer_title_from_body(body),
            meta.get("slug", ""),
            meta.get("type", ""),
            meta.get("status", ""),
            meta.get("price", ""),
            meta.get("regular_price", ""),
            meta.get("sale_price", ""),
            meta.get("stock_status", ""),
            meta.get("stock_quantity", ""),
            meta.get("weight", ""),
            d_len, d_wid, d_hgt,
            sc_name, sc_slug,
            cats_str, tags_str,
            meta.get("product_url", ""),
            img_count,
            wc,
            excerpt,
            fp.name,
        ])

    return headers, rows


def collect_pages(dir_path: Path) -> Tuple[List[str], List[List[Any]]]:
    """
    Return (headers, rows) for a simple page sheet.
    We'll try to use front-matter when present; else infer.
    """
    headers = [
        "title", "slug", "url", "word_count", "excerpt", "file_name"
    ]
    rows: List[List[Any]] = []

    if not dir_path.exists():
        return headers, rows

    files = sorted(dir_path.glob("*.md"))
    for fp in files:
        meta, body = read_md(fp)
        title = meta.get("title", "") or infer_title_from_body(body)
        slug = meta.get("slug", "") or fp.stem
        url = meta.get("url", "") or f"{BASE_URL}/{slug.strip('/')}/"
        plain = markdown_to_plain_text(body)
        wc = word_count(plain)
        excerpt = first_n(plain, 220)

        rows.append([title, slug, url, wc, excerpt, fp.name])

    return headers, rows


def collect_posts() -> Tuple[List[str], List[List[Any]]]:
    """
    Return (headers, rows) for posts.
    """
    headers = [
        "title", "slug", "date", "categories", "tags", "url",
        "word_count", "excerpt", "file_name"
    ]
    rows: List[List[Any]] = []

    if not DIR_POSTS.exists():
        return headers, rows

    files = sorted(DIR_POSTS.glob("*.md"))
    for fp in files:
        meta, body = read_md(fp)
        title = meta.get("title", "") or infer_title_from_body(body)
        slug = meta.get("slug", "") or fp.stem
        date = meta.get("date", "") or meta.get("modified", "")
        url = meta.get("url", "") or f"{BASE_URL}/{slug.strip('/')}/"

        cats = meta.get("categories") or []
        cats_str = ", ".join(cats) if isinstance(cats, list) else str(cats)
        tags = meta.get("tags") or []
        tags_str = ", ".join(tags) if isinstance(tags, list) else str(tags)

        plain = markdown_to_plain_text(body)
        wc = word_count(plain)
        excerpt = first_n(plain, 220)

        rows.append([title, slug, date, cats_str, tags_str, url, wc, excerpt, fp.name])

    return headers, rows

# ---------- main ----------
def main() -> None:
    parser = argparse.ArgumentParser(description="Export pages/posts/products to an Excel workbook.")
    parser.add_argument("--out", default="audit_export.xlsx", help="Output workbook filename (default: audit_export.xlsx)")
    args = parser.parse_args()

    wb = Workbook()
    # openpyxl starts you with a default sheet; reuse it for Products
    ws_products = wb.active
    ws_products.title = "Products"

    # collect and write Products
    p_headers, p_rows = collect_products()
    ensure_sheet_headers(ws_products, p_headers)
    for row in p_rows:
        ws_products.append(row)

    # Pages
    ws_pages = wb.create_sheet("Pages")
    pg_headers, pg_rows = collect_pages(DIR_PAGES)
    ensure_sheet_headers(ws_pages, pg_headers)
    for row in pg_rows:
        ws_pages.append(row)

    # Posts
    ws_posts = wb.create_sheet("Posts")
    post_headers, post_rows = collect_posts()
    ensure_sheet_headers(ws_posts, post_headers)
    for row in post_rows:
        ws_posts.append(row)

    # freeze headers
    for ws in (ws_products, ws_pages, ws_posts):
        ws.freeze_panes = "A2"

    out_path = ROOT / args.out
    wb.save(out_path)
    print(f"✓ Workbook written → {out_path}")

if __name__ == "__main__":
    main()
